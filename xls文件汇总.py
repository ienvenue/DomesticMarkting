import pandas as pd
import openpyxl
import os
import time
# 开始
print('开始汇总.......')
# 文件路径
file_path = r'D:\汇总'
file = r'D:\汇总\汇总.xlsx'
如果存在文件删除
if  os.path.exists(file):
    os.remove(file)

os.chdir(file_path)
file_lst = [file for file in os.listdir()]

wb = openpyxl.load_workbook(file,read_only=True)
ws = openpyxl.load_workbook(file,read_only=True)
sheetName = "Sheet1"
ws = wb[sheetName]

def write_excel(data, file):
    file_name=os.path.basename(file)
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    data.to_excel(writer, file_name.split('-', 1)[0], index=False,merge_cells=True)

for file in file_lst:
    frame = pd.read_excel(file, header=0, sheet_name='Sheet1')
    write_excel(frame,file)

writer.save()
print('汇总完毕，5秒后关闭')

time.sleep(5)